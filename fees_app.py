import os
import sys
import shutil
import sqlite3
import webbrowser
import subprocess
import urllib.parse
import textwrap
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

import openpyxl
from reportlab.lib.pagesizes import A5, A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- PERMANENT PATH SETUP ---
APP_DIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "IntechKidsPlaySchool")
BACKUP_DIR = os.path.join(APP_DIR, "backups")
BILLS_DIR = os.path.join(APP_DIR, "bills")
RECEIPTS_DIR = os.path.join(APP_DIR, "receipts")
REPORTS_DIR = os.path.join(APP_DIR, "reports")
FORMS_DIR = os.path.join(APP_DIR, "admission_forms")

for folder in [APP_DIR, BACKUP_DIR, BILLS_DIR, RECEIPTS_DIR, REPORTS_DIR, FORMS_DIR]:
    os.makedirs(folder, exist_ok=True)

DB_PATH = os.path.join(APP_DIR, "school_erp.db")

# --- UI THEME CONSTANTS ---
BG_MAIN = "#09090B"
BG_SIDEBAR = "#121214"
BG_CARD = "#18181B"
BORDER_COLOR = "#27272A"
FG_PRIMARY = "#FAFAFA"
FG_MUTED = "#A1A1AA"
ACCENT_PRI = "#6366F1"
ACCENT_SEC = "#3F3F46"
SUCCESS = "#10B981"
DANGER = "#EF4444"
WARNING = "#F59E0B"
PURPLE_ACC = "#8B5CF6"
WA_GREEN = "#25D366"
FINANCE_ACC = "#14B8A6"

# --- DATABASE LAYER ---
class SchoolDatabase:
    def __init__(self, db_path=DB_PATH):
        self.db_path = db_path
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        self.create_tables()
        self.seed_default_users()
        self.create_backup()

    def create_tables(self):
        self.cursor.execute("CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT, role TEXT)")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS audit_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, user TEXT, action TEXT, details TEXT)")
        self.cursor.execute("CREATE TABLE IF NOT EXISTS expenses (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, category TEXT, amount REAL, description TEXT, recorded_by TEXT)")
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                roll_no TEXT PRIMARY KEY, name TEXT NOT NULL, class_name TEXT NOT NULL,
                father_name TEXT, mother_name TEXT, contact TEXT NOT NULL,
                transport_opted TEXT DEFAULT 'No', ledger_folio_no TEXT,
                monthly_tuition REAL DEFAULT 0, monthly_transport REAL DEFAULT 0
            )
        """)
        
        self.cursor.execute("PRAGMA table_info(students)")
        cols = [info[1] for info in self.cursor.fetchall()]
        
        migrations = {
            'form_no': "TEXT DEFAULT 'N/A'", 'admission_no': "TEXT DEFAULT 'N/A'",
            'academic_year': "TEXT DEFAULT '2026-27'", 'dob': "TEXT DEFAULT 'N/A'",
            'aadhaar_no': "TEXT DEFAULT 'N/A'", 'place_of_birth': "TEXT DEFAULT 'N/A'",
            'state': "TEXT DEFAULT 'N/A'", 'nationality': "TEXT DEFAULT 'Indian'",
            'religion': "TEXT DEFAULT 'N/A'", 'gender': "TEXT DEFAULT 'N/A'",
            'caste': "TEXT DEFAULT 'General'", 'mother_occupation': "TEXT DEFAULT 'N/A'",
            'father_occupation': "TEXT DEFAULT 'N/A'", 'address': "TEXT DEFAULT 'N/A'",
            'pin_code': "TEXT DEFAULT 'N/A'", 'mother_mobile': "TEXT DEFAULT 'N/A'",
            'email_id': "TEXT DEFAULT 'N/A'", 'mother_tongue': "TEXT DEFAULT 'Hindi'",
            'blood_group': "TEXT DEFAULT 'N/A'", 'identification_marks': "TEXT DEFAULT 'N/A'",
            'prev_school': "TEXT DEFAULT 'N/A'", 'prev_board': "TEXT DEFAULT 'N/A'",
            'marks_percent': "TEXT DEFAULT 'N/A'", 'reference': "TEXT DEFAULT 'N/A'",
            'behaviour': "TEXT DEFAULT 'Normal'", 'class_roll_no': "TEXT DEFAULT 'N/A'"
        }
        
        for col_name, col_def in migrations.items():
            if col_name not in cols:
                self.cursor.execute(f"ALTER TABLE students ADD COLUMN {col_name} {col_def}")

        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS bills (
                bill_no TEXT PRIMARY KEY, roll_no TEXT NOT NULL, billing_month TEXT NOT NULL,
                bill_date TEXT NOT NULL, due_date TEXT NOT NULL, total_amount REAL NOT NULL,
                paid_amount REAL DEFAULT 0, status TEXT DEFAULT 'Unpaid', items_json TEXT NOT NULL,
                FOREIGN KEY (roll_no) REFERENCES students (roll_no)
            )
        """)
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS payments (
                payment_id INTEGER PRIMARY KEY AUTOINCREMENT, bill_no TEXT NOT NULL, roll_no TEXT NOT NULL,
                amount_paid REAL NOT NULL, payment_date TEXT NOT NULL, payment_mode TEXT NOT NULL, receipt_no TEXT,
                FOREIGN KEY (bill_no) REFERENCES bills (bill_no), FOREIGN KEY (roll_no) REFERENCES students (roll_no)
            )
        """)
        self.conn.commit()

    def seed_default_users(self):
        self.cursor.execute("SELECT COUNT(*) FROM users")
        if self.cursor.fetchone()[0] == 0:
            self.cursor.execute("INSERT INTO users VALUES ('admin', 'admin123', 'Admin')")
            self.cursor.execute("INSERT INTO users VALUES ('staff', 'staff123', 'Staff')")
            self.conn.commit()

    def log_audit(self, user, action, details):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.cursor.execute("INSERT INTO audit_logs (timestamp, user, action, details) VALUES (?, ?, ?, ?)", (ts, user, action, details))
        self.conn.commit()

    def authenticate(self, username, password):
        self.cursor.execute("SELECT role FROM users WHERE username = ? AND password = ?", (username, password))
        res = self.cursor.fetchone()
        if res:
            self.log_audit(username, "LOGIN", "Successful login.")
            return True, res[0]
        return False, None

    def create_backup(self):
        if os.path.exists(self.db_path):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            try: shutil.copy2(self.db_path, os.path.join(BACKUP_DIR, f"backup_{ts}.db"))
            except Exception: pass

    def close(self):
        try: self.conn.close()
        except Exception: pass

    def reconnect(self):
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()

    def add_expense(self, category, amount, desc, user):
        date_str = datetime.now().strftime("%Y-%m-%d")
        self.cursor.execute("INSERT INTO expenses (date, category, amount, description, recorded_by) VALUES (?, ?, ?, ?, ?)", (date_str, category, amount, desc, user))
        self.conn.commit(); self.create_backup()

    def get_all_expenses(self):
        self.cursor.execute("SELECT id, date, category, amount, description, recorded_by FROM expenses ORDER BY id DESC")
        return self.cursor.fetchall()

    def promote_all_students(self, user):
        promo_map = {"Play": "Nursery", "Nursery": "LKG", "LKG": "UKG", "UKG": "Class 1", "Class 1": "Class 2", "Class 2": "Class 3", "Class 3": "Class 4", "Class 4": "Class 5"}
        students = self.get_all_students()
        promoted_count = 0
        for st in students:
            curr_class = st[3]
            if curr_class in promo_map:
                self.cursor.execute("UPDATE students SET class_name = ? WHERE roll_no = ?", (promo_map[curr_class], st[0]))
                promoted_count += 1
        self.conn.commit()
        self.log_audit(user, "BULK_PROMOTION", f"Promoted {promoted_count} students.")
        self.create_backup()
        return promoted_count

    def add_student(self, data):
        self.cursor.execute("""
            INSERT INTO students (
                roll_no, form_no, admission_no, academic_year, class_roll_no, name, class_name, 
                dob, aadhaar_no, place_of_birth, state, nationality, religion, gender, caste, 
                father_name, father_occupation, mother_name, mother_occupation, address, pin_code, 
                contact, mother_mobile, email_id, mother_tongue, blood_group, identification_marks, 
                prev_school, prev_board, marks_percent, reference, behaviour, transport_opted, 
                ledger_folio_no, monthly_tuition, monthly_transport
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, data)
        self.conn.commit(); self.create_backup()

    def update_student(self, data):
        self.cursor.execute("""
            UPDATE students SET 
                form_no=?, admission_no=?, academic_year=?, class_roll_no=?, name=?, class_name=?, 
                dob=?, aadhaar_no=?, place_of_birth=?, state=?, nationality=?, religion=?, gender=?, caste=?, 
                father_name=?, father_occupation=?, mother_name=?, mother_occupation=?, address=?, pin_code=?, 
                contact=?, mother_mobile=?, email_id=?, mother_tongue=?, blood_group=?, identification_marks=?, 
                prev_school=?, prev_board=?, marks_percent=?, reference=?, behaviour=?, transport_opted=?, 
                ledger_folio_no=?, monthly_tuition=?, monthly_transport=? 
            WHERE roll_no=?
        """, data)
        self.conn.commit(); self.create_backup()

    def delete_student(self, roll_no, user):
        st = self.get_student(roll_no)
        self.cursor.execute("DELETE FROM payments WHERE roll_no = ?", (roll_no,))
        self.cursor.execute("DELETE FROM bills WHERE roll_no = ?", (roll_no,))
        self.cursor.execute("DELETE FROM students WHERE roll_no = ?", (roll_no,))
        self.conn.commit()
        self.log_audit(user, "DELETE_STUDENT", f"Deleted Enrollment No {roll_no} ({st[2] if st else 'Unknown'})")
        self.create_backup()

    def get_student(self, enrollment_no):
        self.cursor.execute("""
            SELECT roll_no, form_no, admission_no, academic_year, class_roll_no, name, class_name, 
                   dob, aadhaar_no, place_of_birth, state, nationality, religion, gender, caste, 
                   father_name, father_occupation, mother_name, mother_occupation, address, pin_code, 
                   contact, mother_mobile, email_id, mother_tongue, blood_group, identification_marks, 
                   prev_school, prev_board, marks_percent, reference, behaviour, transport_opted, 
                   ledger_folio_no, monthly_tuition, monthly_transport 
            FROM students WHERE roll_no = ?
        """, (enrollment_no,))
        return self.cursor.fetchone()

    def get_all_students(self, search=""):
        q = f"%{search}%"
        self.cursor.execute("""
            SELECT roll_no, class_roll_no, name, class_name, dob, father_name, contact, address 
            FROM students WHERE roll_no LIKE ? OR class_roll_no LIKE ? OR name LIKE ? OR contact LIKE ? 
            ORDER BY roll_no ASC
        """, (q, q, q, q))
        return self.cursor.fetchall()

    def create_bill(self, bill_no, roll_no, month, bill_date, due_date, total, items_str):
        self.cursor.execute("INSERT INTO bills (bill_no, roll_no, billing_month, bill_date, due_date, total_amount, paid_amount, status, items_json) VALUES (?, ?, ?, ?, ?, ?, 0, 'Unpaid', ?)", (bill_no, roll_no, month, bill_date, due_date, total, items_str))
        self.conn.commit(); self.create_backup()

    def get_all_bills(self, status_filter="All Status", search=""):
        q = f"%{search}%"
        query = "SELECT b.bill_no, b.roll_no, s.name, b.billing_month, b.total_amount, b.paid_amount, (b.total_amount - b.paid_amount) AS dues, b.status, b.due_date FROM bills b JOIN students s ON b.roll_no = s.roll_no WHERE (b.bill_no LIKE ? OR b.roll_no LIKE ? OR s.name LIKE ?)"
        params = [q, q, q]
        if status_filter != "All Status": query += " AND b.status = ?"; params.append(status_filter)
        query += " ORDER BY b.bill_date DESC"
        self.cursor.execute(query, tuple(params))
        return self.cursor.fetchall()

    def get_bill_by_id(self, bill_no):
        self.cursor.execute("SELECT * FROM bills WHERE bill_no = ?", (bill_no,))
        return self.cursor.fetchone()

    def record_payment(self, bill_no, amount, mode, receipt_no):
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        bill = self.get_bill_by_id(bill_no)
        if not bill: return False, "Bill not found.", None
        roll_no, total, curr_paid = bill[1], bill[5], bill[6]
        new_paid = curr_paid + amount
        new_status = "Paid" if new_paid >= total else "Partially Paid"
        remaining_dues = max(0.0, total - new_paid)

        self.cursor.execute("UPDATE bills SET paid_amount = ?, status = ? WHERE bill_no = ?", (new_paid, new_status, bill_no))
        self.cursor.execute("INSERT INTO payments (bill_no, roll_no, amount_paid, payment_date, payment_mode, receipt_no) VALUES (?, ?, ?, ?, ?, ?)", (bill_no, roll_no, amount, date_str, mode, receipt_no))
        self.conn.commit(); self.create_backup()
        return True, "Payment recorded successfully.", {"receipt_no": receipt_no, "date": date_str, "mode": mode, "amount_paid": amount, "remaining_dues": remaining_dues}

    def get_payments_for_bill(self, bill_no):
        self.cursor.execute("SELECT payment_id, payment_date, payment_mode, amount_paid, receipt_no FROM payments WHERE bill_no = ? ORDER BY payment_id DESC", (bill_no,))
        return self.cursor.fetchall()

    def get_student_ledger_history(self, roll_no):
        self.cursor.execute("SELECT billing_month, items_json, total_amount, paid_amount, (total_amount - paid_amount), status, bill_no FROM bills WHERE roll_no = ? ORDER BY bill_date ASC", (roll_no,))
        return self.cursor.fetchall()

    def get_financial_dashboard_metrics(self):
        self.cursor.execute("SELECT COUNT(*) FROM students")
        total_students = self.cursor.fetchone()[0] or 0
        self.cursor.execute("SELECT COALESCE(SUM(total_amount), 0), COALESCE(SUM(paid_amount), 0) FROM bills WHERE status != 'Cancelled'")
        billed, collected = self.cursor.fetchone()
        dues = billed - collected
        self.cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM expenses")
        total_expenses = self.cursor.fetchone()[0]
        return total_students, billed, collected, dues, total_expenses, (collected - total_expenses)


# --- ADMISSION FORM PDF GENERATOR ---
def generate_admission_form_pdf(st):
    roll_no = st[0]; form_no = st[1]; admission_no = st[2]; academic_year = st[3]
    class_roll_no = st[4]; name = st[5]; class_name = st[6]; dob = st[7]
    aadhaar_no = st[8]; place_of_birth = st[9]; state = st[10]; nationality = st[11]
    religion = st[12]; gender = st[13]; caste = st[14]; father_name = st[15]
    father_occupation = st[16]; mother_name = st[17]; mother_occupation = st[18]
    address = st[19]; pin_code = st[20]; contact = st[21]; mother_mobile = st[22]
    email_id = st[23]; mother_tongue = st[24]; blood_group = st[25]
    identification_marks = st[26]; prev_school = st[27]; prev_board = st[28]
    marks_percent = st[29]; reference = st[30]; behaviour = st[31]

    pdf_path = os.path.join(FORMS_DIR, f"AdmissionForm_{roll_no}.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('FormTitle', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=14, alignment=1, textColor=colors.HexColor("#000000"))
    sub_style = ParagraphStyle('FormSub', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=9, alignment=1, textColor=colors.HexColor("#FFFFFF"))
    cell_style = ParagraphStyle('FormCell', parent=styles['Normal'], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#000000"))
    cell_bold = ParagraphStyle('FormCellB', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=8.5, leading=11, textColor=colors.HexColor("#000000"))

    elements = []
    elements.append(Paragraph("<b>ADMISSION FORM</b>", title_style))
    elements.append(Spacer(1, 6))

    meta_data = [
        [Paragraph(f"<b>Form No.:</b> {form_no}", cell_style), Paragraph(f"<b>Admission No.:</b> {admission_no}", cell_style), Paragraph("<b>Paste Your Photo Here</b>", ParagraphStyle('PhotoBox', parent=cell_style, alignment=1, fontSize=8, textColor=colors.HexColor("#555555")))],
        [Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d-%m-%Y')}", cell_style), Paragraph(f"<b>Academic Year:</b> {academic_year}", cell_style), ""]
    ]
    meta_table = Table(meta_data, colWidths=[150, 150, 134])
    meta_table.setStyle(TableStyle([('BOX', (0,0), (1,-1), 1, colors.black), ('BOX', (2,0), (2,-1), 1, colors.black), ('SPAN', (2,0), (2,1)), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (2,0), (2,1), 'CENTER'), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
    elements.append(meta_table)
    elements.append(Spacer(1, 4))

    inst_table = Table([[Paragraph("<b>INTECH KIDS PLAY SCHOOL / VECTOR FIELD INSTITUTE</b>", sub_style)]], colWidths=[434])
    inst_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#1E293B")), ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
    elements.append(inst_table)
    elements.append(Spacer(1, 6))

    elements.append(Paragraph("<b>STUDENT PROFILE</b>", cell_bold))
    elements.append(Spacer(1, 4))

    profile_rows = [
        [Paragraph("<b>Name of Pupil (In capital letters):</b>", cell_bold), Paragraph(str(name).upper(), cell_style)],
        [Paragraph("<b>Admission for Class:</b>", cell_bold), Paragraph(f"{class_name}  (Class Roll: {class_roll_no})", cell_style)],
        [Paragraph("<b>Date of Birth:</b>", cell_bold), Paragraph(str(dob), cell_style)],
        [Paragraph("<b>Aadhaar No.:</b>", cell_bold), Paragraph(str(aadhaar_no), cell_style)],
        [Paragraph("<b>Place of Birth / State:</b>", cell_bold), Paragraph(f"{place_of_birth} / {state}", cell_style)],
        [Paragraph("<b>Nationality & Religion:</b>", cell_bold), Paragraph(f"{nationality} / {religion}  (Caste: {caste})", cell_style)],
        [Paragraph("<b>Gender:</b>", cell_bold), Paragraph(str(gender), cell_style)],
        [Paragraph("<b>Mother's Name & Occupation:</b>", cell_bold), Paragraph(f"{mother_name} & {mother_occupation}", cell_style)],
        [Paragraph("<b>Father's Name & Occupation:</b>", cell_bold), Paragraph(f"{father_name} & {father_occupation}", cell_style)],
        [Paragraph("<b>Residential Address & Pin Code:</b>", cell_bold), Paragraph(f"{address} - {pin_code}", cell_style)],
        [Paragraph("<b>Mobile No (Father's / Mother's):</b>", cell_bold), Paragraph(f"Father: {contact} | Mother: {mother_mobile}", cell_style)],
        [Paragraph("<b>E-mail ID:</b>", cell_bold), Paragraph(str(email_id), cell_style)],
        [Paragraph("<b>Mother Tongue & Blood Group:</b>", cell_bold), Paragraph(f"{mother_tongue} | Blood Group: {blood_group}", cell_style)],
        [Paragraph("<b>Identification Marks:</b>", cell_bold), Paragraph(str(identification_marks), cell_style)],
        [Paragraph("<b>Name of Previous School / Board / %:</b>", cell_bold), Paragraph(f"{prev_school} | {prev_board} | {marks_percent}%", cell_style)],
        [Paragraph("<b>Reference / Centre:</b>", cell_bold), Paragraph(str(reference), cell_style)],
        [Paragraph("<b>Appraisal of your child (Achievements):</b>", cell_bold), Paragraph("N/A", cell_style)],
        [Paragraph("<b>General Behaviour:</b>", cell_bold), Paragraph(str(behaviour), cell_style)]
    ]

    prof_table = Table(profile_rows, colWidths=[180, 254])
    prof_table.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 1, colors.black), ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TOPPADDING', (0,0), (-1,-1), 3.5), ('BOTTOMPADDING', (0,0), (-1,-1), 3.5)]))
    elements.append(prof_table)
    elements.append(Spacer(1, 35))

    sign_data = [
        [Paragraph("____________________________<br/><b>(Signature of Parents)</b>", cell_style), Paragraph("____________________________<br/><b>(Signature of Student)</b>", ParagraphStyle('AlignR', parent=cell_style, alignment=2))]
    ]
    sign_table = Table(sign_data, colWidths=[217, 217])
    sign_table.setStyle(TableStyle([('ALIGN', (0,0), (0,0), 'LEFT'), ('ALIGN', (1,0), (1,0), 'RIGHT')]))
    elements.append(sign_table)
    doc.build(elements)
    return pdf_path

# --- SMART CLIPPED ID CARD GENERATOR ---
def generate_advanced_id_cards(students, logo_path="", photos_dir="", sign_path=""):
    pdf_path = os.path.join(REPORTS_DIR, f"ID_Cards_Batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    
    c = canvas.Canvas(pdf_path, pagesize=landscape(A4))
    a4_w, a4_h = landscape(A4)
    card_w, card_h = 2.125 * inch, 3.375 * inch
    cols, rows = 5, 2
    space_x = (a4_w - (cols * card_w)) / (cols + 1)
    space_y = (a4_h - (rows * card_h)) / (rows + 1)

    def draw_card(x, y, st):
        enr_no = st[0]
        class_roll_no = st[4]
        name = st[5]
        class_name = st[6]
        dob = st[7]
        father_name = st[15]
        contact = st[21]
        address = st[19]
        
        c.saveState()
        path = c.beginPath(); path.roundRect(x, y, card_w, card_h, 8); c.clipPath(path, stroke=0)
        c.setFillColor(colors.white); c.rect(x, y, card_w, card_h, stroke=0, fill=1)

        if logo_path and os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, x + 20, y + 40, width=card_w-40, height=card_w-40, preserveAspectRatio=True, mask='auto')
                c.saveState(); c.setFillAlpha(0.88); c.setFillColor(colors.white); c.rect(x, y, card_w, card_h, fill=1, stroke=0); c.restoreState()
            except: pass

        c.setFillColor(colors.HexColor("#F1C40F")); c.ellipse(x - 50, y + card_h - 130, x + card_w + 50, y + card_h + 50, stroke=0, fill=1)
        c.setFillColor(colors.HexColor("#0F2850")); c.ellipse(x - 30, y + card_h - 105, x + card_w + 80, y + card_h + 80, stroke=0, fill=1)
        c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 10.5); c.drawCentredString(x + card_w/2, y + card_h - 18, "INTECH KIDS PLAY SCHOOL")

        addr_w = 130
        c.setFillColor(colors.HexColor("#DC2626")); c.roundRect(x + card_w/2 - addr_w/2, y + card_h - 31, addr_w, 10, 5, stroke=0, fill=1)
        c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 5.5); c.drawCentredString(x + card_w/2, y + card_h - 29, "Bandhubazar, Sohsarai, Nalanda")
        c.setFillColor(colors.HexColor("#F1C40F")); c.setFont("Helvetica-Bold", 6.5); c.drawCentredString(x + card_w/2, y + card_h - 41, "Mob: +91 9304364405")

        c.setFillColor(colors.HexColor("#3498DB")); c.roundRect(x - 2, y + card_h - 135, 18, 55, 4, stroke=0, fill=1)
        c.saveState(); c.translate(x + 10, y + card_h - 135); c.rotate(90); c.setFillColor(colors.white); c.setFont("Helvetica-Bold", 7.5); c.drawCentredString(27.5, 0, "I D   C A R D"); c.restoreState()

        cx, cy, r = x + card_w/2, y + card_h - 85, 28 
        c.setFillColor(colors.white); c.circle(cx, cy, r+2, stroke=0, fill=1)
        
        photo_file = os.path.join(photos_dir, f"{enr_no}.jpg") if photos_dir else ""
        if not os.path.exists(photo_file): photo_file = os.path.join(photos_dir, f"{enr_no}.png") if photos_dir else ""

        c.saveState()
        p = c.beginPath(); p.circle(cx, cy, r); c.clipPath(p, stroke=0)
        if photo_file and os.path.exists(photo_file):
            try: c.drawImage(photo_file, cx - r, cy - r, width=r*2, height=r*2, preserveAspectRatio=False)
            except: c.rect(cx - r, cy - r, r*2, r*2, fill=0)
        else:
            c.setFillColor(colors.HexColor("#F8FAFC")); c.rect(cx - r, cy - r, r*2, r*2, stroke=0, fill=1)
            c.setFillColor(colors.HexColor("#94A3B8")); c.setFont("Helvetica", 6); c.drawCentredString(cx, cy - 2, "PHOTO")
        c.restoreState()

        c.setStrokeColor(colors.HexColor("#CBD5E1")); c.setLineWidth(0.5); c.circle(cx, cy, r, stroke=1, fill=0)
        c.setFillColor(colors.HexColor("#1A3860")); c.setFont("Helvetica-Bold", 12.5); c.drawCentredString(x + card_w/2, cy - r - 14, str(name).title()[:20])

        dy, lh, lbl_x, col_x, val_x = cy - r - 26, 9.5, x + 12, x + 56, x + 60
        c.setFillColor(colors.HexColor("#1A3860")); c.setFont("Helvetica-Bold", 7)
        c.drawString(lbl_x, dy, "Father's Name"); c.drawString(lbl_x, dy - lh, "Class"); c.drawString(lbl_x, dy - lh*2, "Roll No")
        c.drawString(lbl_x, dy - lh*3, "D.O.B."); c.drawString(lbl_x, dy - lh*4, "Mobile"); c.drawString(lbl_x, dy - lh*5, "Address")

        c.setFillColor(colors.HexColor("#16A085")); c.setFont("Helvetica-Bold", 7)
        for i in range(6): c.drawString(col_x, dy - lh*i, ":")

        c.drawString(val_x, dy, str(father_name or 'N/A')[:18].title())
        c.drawString(val_x, dy - lh, str(class_name))
        c.drawString(val_x, dy - lh*2, str(class_roll_no))
        c.drawString(val_x, dy - lh*3, str(dob))
        c.drawString(val_x, dy - lh*4, str(contact))
        
        addr_text = str(address).strip()
        if not addr_text or addr_text.lower() == 'n/a':
            c.drawString(val_x, dy - lh*5, "N/A")
        else:
            wrapped_addr = textwrap.wrap(addr_text, width=22)
            for idx, line in enumerate(wrapped_addr[:3]): c.drawString(val_x, dy - lh*5 - (idx * 8), line.title())

        c.setFillColor(colors.HexColor("#F1C40F")); c.ellipse(x - 20, y - 25, x + card_w + 50, y + 35, stroke=0, fill=1)
        c.setFillColor(colors.HexColor("#0F2850")); c.ellipse(x - 40, y - 35, x + card_w + 20, y + 20, stroke=0, fill=1)

        qr_sz, qr_x, qr_y = 26, x + card_w - 26 - 6, y + 6
        qr_data = f"Name: {name}\nRoll: {class_roll_no}\nClass: {class_name}\nDOB: {dob}\nPh: {contact}\nSchool: INTECH KIDS"
        qr_widget = qr.QrCodeWidget(qr_data); qr_widget.barWidth, qr_widget.barHeight = qr_sz, qr_sz
        d = Drawing(qr_sz, qr_sz); d.add(qr_widget)
        
        c.setFillColor(colors.white); c.roundRect(qr_x - 2, qr_y - 2, qr_sz + 4, qr_sz + 4, 2, stroke=0, fill=1); d.drawOn(c, qr_x, qr_y)

        sig_text_x = x + card_w/3 + 5
        c.setFillColor(colors.white); c.setFont("Helvetica-Oblique", 5.5); c.drawCentredString(sig_text_x, y + 6, "Principal Signature")

        if sign_path and os.path.exists(sign_path):
            try: c.drawImage(sign_path, sig_text_x - 25, y + 12, width=50, height=20, preserveAspectRatio=True, mask=[150, 255, 150, 255, 150, 255])
            except: pass

        c.restoreState(); c.setStrokeColor(colors.HexColor("#94A3B8")); c.setLineWidth(1); c.roundRect(x, y, card_w, card_h, 8, stroke=1, fill=0)

    count = 0
    for st in students:
        col, row = count % cols, (count // cols) % rows
        card_x, card_y = space_x + (col * (card_w + space_x)), a4_h - space_y - card_h - (row * (card_h + space_y))
        draw_card(card_x, card_y, st)
        count += 1
        if count % (cols * rows) == 0 and count < len(students): c.showPage() 
    c.save()
    return pdf_path


# --- STANDARD PDF GENERATORS ---
def get_pdf_styles():
    styles = getSampleStyleSheet()
    header_title = ParagraphStyle('HTitle', alignment=1, fontSize=14, fontName="Helvetica-Bold", leading=16, textColor=colors.HexColor("#09090B"))
    sub_title = ParagraphStyle('HSub', alignment=1, fontSize=8, leading=10, textColor=colors.HexColor("#52525B"))
    tag_style_red = ParagraphStyle('TagR', alignment=1, fontSize=11, fontName="Helvetica-Bold", leading=14, textColor=colors.HexColor("#DC2626"))
    tag_style_green = ParagraphStyle('TagG', alignment=1, fontSize=11, fontName="Helvetica-Bold", leading=14, textColor=colors.HexColor("#059669"))
    b_bold = ParagraphStyle('BBold', fontSize=8, fontName="Helvetica-Bold", leading=11, textColor=colors.HexColor("#18181B"))
    b_norm = ParagraphStyle('BNorm', fontSize=8, leading=11, textColor=colors.HexColor("#27272A"))
    return styles, header_title, sub_title, tag_style_red, tag_style_green, b_bold, b_norm

def create_pdf_header(elements, styles, is_receipt=False):
    _, header_title, sub_title, tag_style_red, tag_style_green, _, _ = styles
    elements.append(Paragraph("<b>INTECH KIDS PLAY SCHOOL</b>", header_title))
    elements.append(Paragraph("Add: Bandhu Bazar, Sohsarai, (Nalanda) 803118<br/>Mob: 9304364405, 9470445172", sub_title))
    elements.append(Spacer(1, 6))
    if is_receipt: elements.append(Paragraph("<b>FEE PAYMENT RECEIPT</b>", tag_style_green))
    else: elements.append(Paragraph("<b>DEMAND BILL / FEE INVOICE</b>", tag_style_red))
    elements.append(Spacer(1, 8))

def generate_demand_bill_pdf(bill_data, student_data, items_list):
    bill_no, _, month, bill_date, due_date, total_amt, paid_amt, status, _ = bill_data
    enr_no = student_data[0]
    name = student_data[5]
    class_name = student_data[6]
    f_name = student_data[15]
    contact = student_data[21]

    pdf_path = os.path.join(BILLS_DIR, f"DemandBill_{bill_no}.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=A5, rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    styles = get_pdf_styles()
    _, _, _, _, _, b_bold, b_norm = styles
    
    elements = []
    create_pdf_header(elements, styles, is_receipt=False)

    meta_info = [
        [Paragraph(f"<b>Bill No:</b> {bill_no}", b_bold), Paragraph(f"<b>Date:</b> {bill_date}", b_norm), Paragraph(f"<b>Month:</b> {month}", b_bold)],
        [Paragraph(f"<b>Name:</b> {name}", b_bold), Paragraph(f"<b>Enr No:</b> {enr_no}", b_norm), Paragraph(f"<b>Class:</b> {class_name}", b_norm)],
        [Paragraph(f"<b>Father:</b> {f_name or 'N/A'}", b_norm), Paragraph(f"<b>Contact:</b> {contact}", b_norm), Paragraph(f"<b>Due Date:</b> {due_date}", b_bold)]
    ]
    meta_table = Table(meta_info, colWidths=[140, 130, 110])
    meta_table.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#D4D4D8")), ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F4F4F5")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
    elements.append(meta_table); elements.append(Spacer(1, 10))

    item_rows = [[Paragraph("<b>Sl. No.</b>", b_bold), Paragraph("<b>Particulars</b>", b_bold), Paragraph("<b>Amount (Rs.)</b>", b_bold)]]
    for idx, (particular, amount) in enumerate(items_list, start=1):
        if amount > 0: item_rows.append([Paragraph(str(idx), b_norm), Paragraph(particular, b_norm), Paragraph(f"{amount:,.2f}", b_norm)])
    item_rows.append([Paragraph("<b>-</b>", b_bold), Paragraph("<b>Gross Total</b>", b_bold), Paragraph(f"<b>Rs. {total_amt:,.2f}</b>", b_bold)])
    item_rows.append([Paragraph("<b>-</b>", b_bold), Paragraph("<b>Paid Amount</b>", b_bold), Paragraph(f"Rs. {paid_amt:,.2f}", b_norm)])
    item_rows.append([Paragraph("<b>-</b>", b_bold), Paragraph("<b>Balance Dues</b>", b_bold), Paragraph(f"<b>Rs. {max(0.0, total_amt - paid_amt):,.2f}</b>", b_bold)])
    
    item_table = Table(item_rows, colWidths=[40, 240, 100])
    item_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D4D4D8")), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E4E4E7")), ('ALIGN', (2,0), (2,-1), 'RIGHT'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
    elements.append(item_table); elements.append(Spacer(1, 30))

    sign_table = Table([["", ""], ["Authorized Signature", "Parent's Signature"]], colWidths=[190, 190])
    sign_table.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('LINEABOVE', (0,1), (0,1), 0.8, colors.HexColor("#71717A")), ('LINEABOVE', (1,1), (1,1), 0.8, colors.HexColor("#71717A")), ('FONTSIZE', (0,1), (-1,1), 8)]))
    elements.append(sign_table); doc.build(elements)
    return pdf_path

def generate_receipt_pdf(payment_data, bill_data, student_data):
    bill_no, _, month, _, _, total_amt, _, _, _ = bill_data
    enr_no = student_data[0]
    name = student_data[5]
    class_name = student_data[6]

    rec_no = payment_data["receipt_no"]
    pdf_path = os.path.join(RECEIPTS_DIR, f"Receipt_{rec_no}.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=A5, rightMargin=15, leftMargin=15, topMargin=15, bottomMargin=15)
    styles = get_pdf_styles()
    _, _, _, _, _, b_bold, b_norm = styles
    elements = []
    create_pdf_header(elements, styles, is_receipt=True)

    meta_info = [
        [Paragraph(f"<b>Receipt No:</b> {rec_no}", b_bold), Paragraph(f"<b>Date:</b> {payment_data['date']}", b_norm)],
        [Paragraph(f"<b>Against Bill:</b> {bill_no} ({month})", b_norm), Paragraph(f"<b>Mode:</b> {payment_data['mode']}", b_bold)],
        [Paragraph(f"<b>Student Name:</b> {name}", b_bold), Paragraph(f"<b>Enr No:</b> {enr_no} | Class: {class_name}", b_norm)],
    ]
    meta_table = Table(meta_info, colWidths=[200, 180])
    meta_table.setStyle(TableStyle([('BOX', (0,0), (-1,-1), 0.5, colors.HexColor("#D4D4D8")), ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F4F4F5")), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
    elements.append(meta_table); elements.append(Spacer(1, 15))
    
    pay_data = [
        [Paragraph("<b>Payment Details</b>", b_bold), Paragraph("<b>Amount (Rs.)</b>", b_bold)],
        [Paragraph("Total Bill Amount", b_norm), Paragraph(f"{total_amt:,.2f}", b_norm)],
        [Paragraph("<b>Amount Paid Now</b>", b_bold), Paragraph(f"<b>{payment_data['amount_paid']:,.2f}</b>", b_bold)],
        [Paragraph("<b>Remaining Dues</b>", b_bold), Paragraph(f"<b>{payment_data['remaining_dues']:,.2f}</b>", b_bold)]
    ]
    pay_table = Table(pay_data, colWidths=[280, 100])
    pay_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D4D4D8")), ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E4E4E7")), ('BACKGROUND', (0,2), (-1,2), colors.HexColor("#D1FAE5")), ('ALIGN', (1,0), (1,-1), 'RIGHT'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
    elements.append(pay_table); elements.append(Spacer(1, 30))

    sign_table = Table([["", ""], ["Authorized Signature", "Parent's Signature"]], colWidths=[190, 190])
    sign_table.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('LINEABOVE', (0,1), (0,1), 0.8, colors.HexColor("#71717A")), ('LINEABOVE', (1,1), (1,1), 0.8, colors.HexColor("#71717A")), ('FONTSIZE', (0,1), (-1,1), 8)]))
    elements.append(sign_table); doc.build(elements)
    return pdf_path

# --- LOGIN SYSTEM ---
class LoginApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("System Login")
        self.geometry("400x500")
        self.configure(bg=BG_MAIN)
        self.eval('tk::PlaceWindow . center')
        try: self.iconbitmap(resource_path("school_logo.ico"))
        except: pass
        self.db = SchoolDatabase()
        f = tk.Frame(self, bg=BG_CARD, padx=30, pady=30, highlightthickness=1, highlightbackground=BORDER_COLOR); f.pack(expand=True)
        tk.Label(f, text="INTECH KIDS", font=("Segoe UI", 16, "bold"), fg=ACCENT_PRI, bg=BG_CARD).pack(pady=(0, 5))
        tk.Label(f, text="Secure ERP Login", font=("Segoe UI", 10), fg=FG_MUTED, bg=BG_CARD).pack(pady=(0, 25))
        tk.Label(f, text="Username", font=("Segoe UI", 9, "bold"), fg=FG_PRIMARY, bg=BG_CARD).pack(anchor="w")
        self.ent_user = tk.Entry(f, width=25, font=("Segoe UI", 11), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); self.ent_user.pack(pady=(5, 15), ipady=4)
        tk.Label(f, text="Password", font=("Segoe UI", 9, "bold"), fg=FG_PRIMARY, bg=BG_CARD).pack(anchor="w")
        self.ent_pass = tk.Entry(f, width=25, font=("Segoe UI", 11), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid", show="*"); self.ent_pass.pack(pady=(5, 25), ipady=4)
        tk.Button(f, text="LOGIN", font=("Segoe UI", 10, "bold"), bg=ACCENT_PRI, fg="#FFFFFF", bd=0, padx=15, pady=8, cursor="hand2", command=self.do_login).pack(fill="x")
        self.bind('<Return>', lambda e: self.do_login())

    def do_login(self):
        u, p = self.ent_user.get().strip(), self.ent_pass.get().strip()
        success, role = self.db.authenticate(u, p)
        if success:
            self.destroy(); app = SchoolFeeApp(u, role); app.mainloop()
        else: messagebox.showerror("Access Denied", "Invalid username or password.")

# --- MAIN ERP APPLICATION ---
class SchoolFeeApp(tk.Tk):
    def __init__(self, username, role):
        super().__init__()
        self.current_user = username
        self.current_role = role
        self.title(f"Intech Kids Play School - ERP | User: {self.current_user} ({self.current_role})")
        self.geometry("1280x800")
        self.minsize(1100, 700)
        self.configure(bg=BG_MAIN) 
        try: self.iconbitmap(resource_path("school_logo.ico"))
        except: pass
        self.db = SchoolDatabase()
        
        self.id_logo_path = ""
        self.id_photos_dir = ""
        self.id_sign_path = ""

        self.fee_heads = ["Admission Fee", "Monthly / Tuition Fee", "Transport Charge", "Student Kit & Dress", "Annual Dev. Fee", "Exam Charge", "Maintenance Charge", "Computer Fee", "Late Fine", "Back Dues", "Miscellaneous / Other"]
        self.create_theme_styles()
        self.create_sidebar_layout()
        self.show_dashboard()

    def create_theme_styles(self):
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("Treeview", background=BG_SIDEBAR, foreground=FG_PRIMARY, fieldbackground=BG_SIDEBAR, rowheight=35, font=("Segoe UI", 9), borderwidth=0)
        self.style.configure("Treeview.Heading", background=BG_CARD, foreground=FG_MUTED, font=("Segoe UI", 9, "bold"), relief="flat", padding=[8, 8])
        self.style.map("Treeview", background=[("selected", ACCENT_PRI)], foreground=[("selected", "#FFFFFF")])
        self.style.map("Treeview.Heading", background=[("active", BORDER_COLOR)])
        self.style.configure("TCombobox", fieldbackground=BG_MAIN, background=BG_CARD, foreground=FG_PRIMARY, arrowcolor=ACCENT_PRI)
        self.style.configure("Horizontal.TScrollbar", background=BG_CARD, bordercolor=BG_MAIN, arrowcolor=FG_PRIMARY)
        self.style.configure("Vertical.TScrollbar", background=BG_CARD, bordercolor=BG_MAIN, arrowcolor=FG_PRIMARY)

    def create_sidebar_layout(self):
        top_header = tk.Frame(self, bg=BG_MAIN, height=55, padx=20)
        top_header.pack(fill="x", side="top")
        tk.Label(top_header, text="🏫 INTECH KIDS PLAY SCHOOL", font=("Segoe UI", 13, "bold"), fg=ACCENT_PRI, bg=BG_MAIN).pack(side="left", pady=15)
        tk.Label(top_header, text="| Complete ERP System", font=("Segoe UI", 10), fg=FG_MUTED, bg=BG_MAIN).pack(side="left", padx=10, pady=15)
        tk.Label(top_header, text=f"Logged in as: {self.current_user} ({self.current_role})", font=("Segoe UI", 9, "bold"), fg=SUCCESS, bg=BG_MAIN).pack(side="right", pady=15)

        body = tk.Frame(self, bg=BG_MAIN); body.pack(fill="both", expand=True)
        self.sidebar = tk.Frame(body, bg=BG_SIDEBAR, width=230, padx=12, pady=25)
        self.sidebar.pack(side="left", fill="y"); self.sidebar.pack_propagate(False)
        tk.Label(self.sidebar, text="MAIN MENU", font=("Segoe UI", 8, "bold"), fg=FG_MUTED, bg=BG_SIDEBAR).pack(anchor="w", padx=10, pady=(0, 10))

        self.nav_buttons = {}
        nav_items = [
            ("dash", "📊  Dashboard", self.show_dashboard),
            ("students", "👨‍🎓  Students Directory", self.show_students),
            ("register", "➕  Register & Edit", self.show_register),
            ("bills", "💳  Bills & Payments", self.show_bills),
            ("import", "📥  Excel Bulk Import", self.show_import_excel),
            ("id_studio", "🪪  ID Card Studio", self.show_id_studio),
        ]
        if self.current_role == "Admin": nav_items.append(("finance", "💼  Finance & Expenses", self.show_finance))

        for key, text, cmd in nav_items:
            btn = tk.Button(self.sidebar, text=text, font=("Segoe UI", 10, "bold"), anchor="w", padx=15, pady=12, bg=BG_SIDEBAR, fg=FG_MUTED, activebackground=BG_CARD, activeforeground=FG_PRIMARY, bd=0, cursor="hand2", command=cmd)
            btn.pack(fill="x", pady=3); self.nav_buttons[key] = btn

        if self.current_role == "Admin":
            tk.Label(self.sidebar, text="SYSTEM & ADMIN", font=("Segoe UI", 8, "bold"), fg=FG_MUTED, bg=BG_SIDEBAR).pack(anchor="w", padx=10, pady=(30, 10))
            tk.Button(self.sidebar, text="🎓 Promote Students", font=("Segoe UI", 9), anchor="w", padx=15, pady=8, bg=BG_CARD, fg=FG_PRIMARY, bd=0, cursor="hand2", command=self.run_promotion).pack(fill="x", pady=3)
            tk.Button(self.sidebar, text="💾 Restore Database", font=("Segoe UI", 9), anchor="w", padx=15, pady=8, bg=BG_CARD, fg=FG_PRIMARY, bd=0, cursor="hand2", command=self.restore_backup_dialog).pack(fill="x", pady=3)
        
        tk.Button(self.sidebar, text="📂 Open Reports/Bills", font=("Segoe UI", 9), anchor="w", padx=15, pady=8, bg=BG_CARD, fg=FG_PRIMARY, bd=0, cursor="hand2", command=lambda: os.startfile(APP_DIR) if hasattr(os, 'startfile') else webbrowser.open(APP_DIR)).pack(fill="x", pady=3)
        self.content_frame = tk.Frame(body, bg=BG_MAIN, padx=30, pady=25); self.content_frame.pack(side="right", fill="both", expand=True)

    def set_active_nav(self, active_key):
        for key, btn in self.nav_buttons.items():
            if key == active_key: btn.config(bg=ACCENT_PRI, fg="#FFFFFF")
            else: btn.config(bg=BG_SIDEBAR, fg=FG_MUTED)

    def clear_content(self):
        for widget in self.content_frame.winfo_children(): widget.destroy()

    def create_card_frame(self, parent): return tk.Frame(parent, bg=BG_CARD, highlightthickness=1, highlightbackground=BORDER_COLOR)
    def create_btn(self, parent, text, color, cmd, icon=""): return tk.Button(parent, text=f"{icon} {text}" if icon else text, font=("Segoe UI", 9, "bold"), bg=color, fg="#FFFFFF", bd=0, padx=12, pady=7, cursor="hand2", command=cmd)

    # --- VIEW: DASHBOARD ---
    def show_dashboard(self):
        self.set_active_nav("dash"); self.clear_content()
        tk.Label(self.content_frame, text="Financial Dashboard & Net Income", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")
        st_count, billed, collected, dues, expenses, net = self.db.get_financial_dashboard_metrics()
        
        cards_grid1 = tk.Frame(self.content_frame, bg=BG_MAIN); cards_grid1.pack(fill="x", pady=(15, 5))
        self.render_card(cards_grid1, "TOTAL STUDENTS", str(st_count), "Registered", ACCENT_PRI)
        self.render_card(cards_grid1, "TOTAL BILLED", f"₹{billed:,.2f}", "Demand Generated", PURPLE_ACC)
        self.render_card(cards_grid1, "OUTSTANDING DUES", f"₹{dues:,.2f}", "Pending Amount", WARNING)

        cards_grid2 = tk.Frame(self.content_frame, bg=BG_MAIN); cards_grid2.pack(fill="x", pady=5)
        self.render_card(cards_grid2, "TOTAL COLLECTED", f"₹{collected:,.2f}", "Gross Revenue", SUCCESS)
        self.render_card(cards_grid2, "TOTAL EXPENSES", f"₹{expenses:,.2f}", "Salaries, Rent, Misc", DANGER)
        self.render_card(cards_grid2, "NET INCOME", f"₹{net:,.2f}", "Collected minus Expenses", FINANCE_ACC)

        recent_frame = self.create_card_frame(self.content_frame); recent_frame.pack(fill="both", expand=True, pady=15)
        tk.Label(recent_frame, text="Recent Billing Activity", font=("Segoe UI", 12, "bold"), fg=FG_PRIMARY, bg=BG_CARD).pack(anchor="w", padx=20, pady=(15, 5))
        cols = ("bill_no", "enr_no", "name", "total", "status")
        tree = ttk.Treeview(recent_frame, columns=cols, show="headings", height=6)
        for c, h in zip(cols, ["Bill No.", "Enrollment No", "Student Name", "Total (₹)", "Status"]): tree.heading(c, text=h); tree.column(c, anchor="center" if c not in ["name", "total"] else ("w" if c=="name" else "e"))
        for r in self.db.get_all_bills("All Status")[:8]: tree.insert("", "end", values=(r[0], r[1], r[2], f"{r[4]:.2f}", r[7]))
        tree.pack(fill="both", expand=True, padx=20, pady=(0, 20))

    def render_card(self, parent, title, val, subtitle, accent):
        f = tk.Frame(parent, bg=BG_CARD, padx=20, pady=15, highlightthickness=1, highlightbackground=BORDER_COLOR); f.pack(side="left", fill="both", expand=True, padx=8)
        tk.Label(f, text=title, font=("Segoe UI", 8, "bold"), fg=FG_MUTED, bg=BG_CARD).pack(anchor="w")
        tk.Label(f, text=val, font=("Segoe UI", 16, "bold"), fg=accent, bg=BG_CARD).pack(anchor="w", pady=(4, 2))
        tk.Label(f, text=subtitle, font=("Segoe UI", 8), fg=FG_MUTED, bg=BG_CARD).pack(anchor="w")

    # --- VIEW: STUDENTS (ADMISSION DESK) ---
    def show_students(self):
        self.set_active_nav("students"); self.clear_content()
        tk.Label(self.content_frame, text="Student Directory & Admission Desk", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")
        
        bar = self.create_card_frame(self.content_frame); bar.pack(fill="x", pady=15); bar.config(padx=15, pady=12)
        tk.Label(bar, text="Search:", font=("Segoe UI", 9), fg=FG_MUTED, bg=BG_CARD).pack(side="left", padx=5)
        self.ent_st_search = tk.Entry(bar, width=25, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid")
        self.ent_st_search.pack(side="left", padx=5); self.ent_st_search.bind("<KeyRelease>", lambda e: self.populate_students_tree())

        if self.current_role == "Admin": 
            self.create_btn(bar, "Remove", DANGER, self.remove_selected_student, "🗑️").pack(side="right", padx=3)
            self.create_btn(bar, "Edit Record", FINANCE_ACC, self.open_edit_student_dialog, "✏️").pack(side="right", padx=3)
            
        self.create_btn(bar, "Print Admission Form", SUCCESS, self.print_admission_form, "🖨️").pack(side="right", padx=3)
        self.create_btn(bar, "12-Month Ledger", ACCENT_PRI, self.open_digital_fee_card, "📖").pack(side="right", padx=3)

        tbl_frame = tk.Frame(self.content_frame, bg=BG_MAIN); tbl_frame.pack(fill="both", expand=True)
        cols = ("enroll", "roll", "name", "class", "dob", "father", "contact", "address")
        self.tree_st = ttk.Treeview(tbl_frame, columns=cols, show="headings", selectmode="browse")
        w_dict = {"enroll": 90, "roll": 70, "name": 180, "class": 80, "dob": 100, "father": 160, "contact": 110, "address": 200}
        for c, h in zip(cols, ["Enrollment No", "Class Roll", "Student Name", "Class", "D.O.B", "Father's Name", "Contact", "Address"]): self.tree_st.heading(c, text=h); self.tree_st.column(c, width=w_dict[c], minwidth=w_dict[c], anchor="center" if c not in ["name", "father", "address"] else "w")
        
        scroll_y = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree_st.yview)
        scroll_x = ttk.Scrollbar(tbl_frame, orient="horizontal", command=self.tree_st.xview)
        self.tree_st.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_x.pack(side="bottom", fill="x"); self.tree_st.pack(side="left", fill="both", expand=True); scroll_y.pack(side="right", fill="y")
        self.tree_st.bind("<Double-1>", lambda e: self.open_digital_fee_card()); self.populate_students_tree()

    def populate_students_tree(self):
        for item in self.tree_st.get_children(): self.tree_st.delete(item)
        for r in self.db.get_all_students(self.ent_st_search.get().strip()): self.tree_st.insert("", "end", values=r)

    def print_admission_form(self):
        sel = self.tree_st.selection()
        if not sel: return messagebox.showwarning("Select Student", "Please select a student from the table to print their admission form.")
        enr_no = self.tree_st.item(sel[0])["values"][0]
        st = self.db.get_student(str(enr_no))
        if st:
            try:
                pdf_path = generate_admission_form_pdf(st)
                messagebox.showinfo("Success", f"Admission form generated successfully!\nSaved to:\n{pdf_path}")
                if hasattr(os, 'startfile'): os.startfile(pdf_path)
                else: webbrowser.open(pdf_path)
            except Exception as e:
                messagebox.showerror("Error", f"Could not generate admission form: {e}")

    def remove_selected_student(self):
        sel = self.tree_st.selection()
        if not sel: return
        enr_no = self.tree_st.item(sel[0])["values"][0]
        st = self.db.get_student(str(enr_no))
        if messagebox.askyesno("Confirm Delete", f"⚠️ Permanently delete {st[5]} (Enrollment: {enr_no})?\n\nThis deletes all bills and payments. Admin Action will be audited.", icon="warning"):
            self.db.delete_student(str(enr_no), self.current_user)
            messagebox.showinfo("Deleted", "Student removed safely."); self.populate_students_tree()

    def open_digital_fee_card(self):
        sel = self.tree_st.selection()
        if not sel: return
        enr_no = self.tree_st.item(sel[0])["values"][0]
        st = self.db.get_student(str(enr_no)); ledger_rows = self.db.get_student_ledger_history(str(enr_no))

        win = tk.Toplevel(self); win.title(f"Fee Ledger - {st[5]}"); win.geometry("980x560"); win.configure(bg=BG_MAIN)
        h_frame = self.create_card_frame(win); h_frame.pack(fill="x", padx=20, pady=20); h_frame.config(padx=20, pady=15)
        tk.Label(h_frame, text=f"ANNUAL LEDGER: {st[5]} (Enrollment: {st[0]} | Class: {st[6]})", font=("Segoe UI", 12, "bold"), fg=ACCENT_PRI, bg=BG_CARD).pack(anchor="w")
        tk.Label(h_frame, text=f"Father: {st[15]} | Contact: {st[21]}", font=("Segoe UI", 9), fg=FG_MUTED, bg=BG_CARD).pack(anchor="w", pady=(4,0))

        t_frame = tk.Frame(win, bg=BG_MAIN, padx=20, pady=10); t_frame.pack(fill="both", expand=True)
        cols = ("month", "bill", "breakup", "total", "paid", "dues", "status")
        t = ttk.Treeview(t_frame, columns=cols, show="headings")
        w_dict = {"month":90, "bill":140, "breakup":320, "total":90, "paid":90, "dues":90, "status":100}
        for c, h in zip(cols, ["Month", "Bill No.", "Particulars Breakdown", "Total (₹)", "Paid (₹)", "Dues (₹)", "Status"]): t.heading(c, text=h); t.column(c, width=w_dict[c], minwidth=w_dict[c], anchor="e" if c in ["total","paid","dues"] else "center" if c!="breakup" else "w")
        for r in ledger_rows: t.insert("", "end", values=(r[0], r[6], r[1].replace(", ", " | "), f"{r[2]:.2f}", f"{r[3]:.2f}", f"{r[4]:.2f}", r[5]))
        t.pack(fill="both", expand=True)

    def open_edit_student_dialog(self):
        sel = self.tree_st.selection()
        if not sel: return messagebox.showwarning("Select Student", "Please select a student from the table to edit.")
        enr_no = self.tree_st.item(sel[0])["values"][0]
        st = self.db.get_student(str(enr_no))
        
        win = tk.Toplevel(self); win.title(f"Edit Student Profile - {st[5]}"); win.geometry("900x700"); win.configure(bg=BG_MAIN)
        
        canvas_container = tk.Canvas(win, bg=BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas_container.yview)
        card = tk.Frame(canvas_container, bg=BG_CARD, padx=30, pady=25)

        canvas_window = canvas_container.create_window((0, 0), window=card, anchor="nw")

        def configure_card(event):
            canvas_container.configure(scrollregion=canvas_container.bbox("all"))
        def configure_win_canvas(event):
            canvas_container.itemconfig(canvas_window, width=event.width)

        card.bind("<Configure>", configure_card)
        canvas_container.bind("<Configure>", configure_win_canvas)
        canvas_container.configure(yscrollcommand=scrollbar.set)

        canvas_container.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        scrollbar.pack(side="right", fill="y", pady=20)
        
        tk.Label(card, text="Edit Admission Profile", font=("Segoe UI", 16, "bold"), fg=FINANCE_ACC, bg=BG_CARD).grid(row=0, column=0, columnspan=4, sticky="w", padx=10, pady=(0, 15))

        labels = [
            ("Enrollment No *:", 1, 0, st[0]), ("Form No.:", 1, 2, st[1]),
            ("Admission No.:", 2, 0, st[2]), ("Academic Year:", 2, 2, st[3]),
            ("Class Roll No *:", 3, 0, st[4]), ("Student Name *:", 3, 2, st[5]),
            ("Class / Grade *:", 4, 0, st[6]), ("D.O.B (DD-MM-YYYY) *:", 4, 2, st[7]),
            ("Aadhaar No.:", 5, 0, st[8]), ("Place of Birth:", 5, 2, st[9]),
            ("State:", 6, 0, st[10]), ("Nationality:", 6, 2, st[11]),
            ("Religion:", 7, 0, st[12]), ("Gender:", 7, 2, st[13]),
            ("Caste (General/SC/ST/OBC):", 8, 0, st[14]), ("Father's Name:", 8, 2, st[15]),
            ("Father's Occupation:", 9, 0, st[16]), ("Mother's Name:", 9, 2, st[17]),
            ("Mother's Occupation:", 10, 0, st[18]), ("Residential Address:", 10, 2, st[19]),
            ("Pin Code:", 11, 0, st[20]), ("Mobile (Father's):", 11, 2, st[21]),
            ("Mobile (Mother's):", 12, 0, st[22]), ("E-mail ID:", 12, 2, st[23]),
            ("Mother Tongue:", 13, 0, st[24]), ("Blood Group:", 13, 2, st[25]),
            ("Identification Marks:", 14, 0, st[26]), ("Name of Previous School:", 14, 2, st[27]),
            ("Name of Board:", 15, 0, st[28]), ("% of Marks:", 15, 2, st[29]),
            ("Reference / Centre:", 16, 0, st[30]), ("Behaviour:", 16, 2, st[31]),
            ("Transport Opted (Yes/No):", 17, 0, st[32]), ("Ledger Folio No.:", 17, 2, st[33]),
            ("Base Tuition Fee (₹):", 18, 0, st[34]), ("Base Transport Fee (₹):", 18, 2, st[35])
        ]
        
        edit_entries = {}
        for text, r, c, val in labels:
            tk.Label(card, text=text, font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=r, column=c, sticky="w", padx=10, pady=5)
            if "Transport Opted" in text:
                cb = ttk.Combobox(card, values=["No", "Yes"], state="readonly", width=22)
                cb.set(val)
                cb.grid(row=r, column=c+1, padx=10, pady=5)
                edit_entries[text] = cb
            else:
                e = tk.Entry(card, width=22, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid")
                e.insert(0, str(val if val is not None else ""))
                e.grid(row=r, column=c+1, padx=10, pady=5)
                edit_entries[text] = e
                if "Enrollment No" in text: e.config(state="readonly", fg=FG_MUTED)

        def save_edit():
            enr = edit_entries["Enrollment No *:"].get().strip()
            name = edit_entries["Student Name *:"].get().strip()
            cls = edit_entries["Class / Grade *:"].get().strip()
            dob = edit_entries["D.O.B (DD-MM-YYYY) *:"].get().strip()
            contact = edit_entries["Mobile (Father's):"].get().strip()
            
            if not (enr and name and cls and contact and dob): return messagebox.showerror("Error", "Required fields cannot be empty.")
            
            try: tuition, tr_fee = float(edit_entries["Base Tuition Fee (₹):"].get().strip() or 0), float(edit_entries["Base Transport Fee (₹):"].get().strip() or 0)
            except ValueError: return messagebox.showerror("Error", "Fees must be numeric.")

            data = (
                edit_entries["Form No.:"].get().strip(),
                edit_entries["Admission No.:"].get().strip(),
                edit_entries["Academic Year:"].get().strip(),
                edit_entries["Class Roll No *:"].get().strip(),
                name, cls, dob,
                edit_entries["Aadhaar No.:"].get().strip(),
                edit_entries["Place of Birth:"].get().strip(),
                edit_entries["State:"].get().strip(),
                edit_entries["Nationality:"].get().strip(),
                edit_entries["Religion:"].get().strip(),
                edit_entries["Gender:"].get().strip(),
                edit_entries["Caste (General/SC/ST/OBC):"].get().strip(),
                edit_entries["Father's Name:"].get().strip(),
                edit_entries["Father's Occupation:"].get().strip(),
                edit_entries["Mother's Name:"].get().strip(),
                edit_entries["Mother's Occupation:"].get().strip(),
                edit_entries["Residential Address:"].get().strip(),
                edit_entries["Pin Code:"].get().strip(),
                contact,
                edit_entries["Mobile (Mother's):"].get().strip(),
                edit_entries["E-mail ID:"].get().strip(),
                edit_entries["Mother Tongue:"].get().strip(),
                edit_entries["Blood Group:"].get().strip(),
                edit_entries["Identification Marks:"].get().strip(),
                edit_entries["Name of Previous School:"].get().strip(),
                edit_entries["Name of Board:"].get().strip(),
                edit_entries["% of Marks:"].get().strip(),
                edit_entries["Reference / Centre:"].get().strip(),
                edit_entries["Behaviour:"].get().strip(),
                edit_entries["Transport Opted (Yes/No):"].get().strip(),
                edit_entries["Ledger Folio No.:"].get().strip(),
                tuition, tr_fee,
                enr
            )
            try: 
                self.db.update_student(data)
                messagebox.showinfo("Success", f"'{name}' admission profile updated successfully!")
                self.populate_students_tree(); win.destroy()
            except Exception as e: messagebox.showerror("Error", f"Failed to update: {e}")

        self.create_btn(card, "Update Full Admission Profile", FINANCE_ACC, save_edit, "💾").grid(row=19, column=0, columnspan=4, pady=(15, 0))

    # --- VIEW: ID CARD STUDIO ---
    def show_id_studio(self):
        self.set_active_nav("id_studio"); self.clear_content()
        tk.Label(self.content_frame, text="🪪 ID Card Studio", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")
        tk.Label(self.content_frame, text="Generate perfect A4 10-grid standard ID Cards matching the design.", font=("Segoe UI", 9), fg=FG_MUTED, bg=BG_MAIN).pack(anchor="w", pady=(2, 15))

        studio = tk.Frame(self.content_frame, bg=BG_MAIN); studio.pack(fill="both", expand=True)
        
        left_f = self.create_card_frame(studio); left_f.pack(side="left", fill="both", expand=True, padx=(0,10)); left_f.config(padx=20, pady=20)
        tk.Label(left_f, text="Design & Media Settings", font=("Segoe UI", 11, "bold"), fg=ACCENT_PRI, bg=BG_CARD).pack(anchor="w", pady=(0, 15))

        tk.Label(left_f, text="School Logo Image:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).pack(anchor="w", pady=5)
        logo_frame = tk.Frame(left_f, bg=BG_CARD); logo_frame.pack(fill="x", pady=2)
        self.lbl_logo = tk.Label(logo_frame, text=os.path.basename(self.id_logo_path) if self.id_logo_path else "No file selected", width=30, bg=BG_MAIN, fg=FG_PRIMARY, bd=1, relief="solid", anchor="w", padx=5)
        self.lbl_logo.pack(side="left", ipady=3)
        self.create_btn(logo_frame, "Browse", ACCENT_SEC, self.pick_logo).pack(side="left", padx=5)

        tk.Label(left_f, text="Principal Signature (Optional):", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).pack(anchor="w", pady=(15, 5))
        sign_frame = tk.Frame(left_f, bg=BG_CARD); sign_frame.pack(fill="x", pady=2)
        self.lbl_sign = tk.Label(sign_frame, text=os.path.basename(self.id_sign_path) if self.id_sign_path else "No file selected", width=30, bg=BG_MAIN, fg=FG_PRIMARY, bd=1, relief="solid", anchor="w", padx=5)
        self.lbl_sign.pack(side="left", ipady=3)
        self.create_btn(sign_frame, "Browse", ACCENT_SEC, self.pick_sign).pack(side="left", padx=5)

        tk.Label(left_f, text="Student Photos Directory:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).pack(anchor="w", pady=(15, 5))
        photo_frame = tk.Frame(left_f, bg=BG_CARD); photo_frame.pack(fill="x", pady=2)
        self.lbl_photos = tk.Label(photo_frame, text=os.path.basename(self.id_photos_dir) if self.id_photos_dir else "No folder selected", width=30, bg=BG_MAIN, fg=FG_PRIMARY, bd=1, relief="solid", anchor="w", padx=5)
        self.lbl_photos.pack(side="left", ipady=3)
        self.create_btn(photo_frame, "Browse", ACCENT_SEC, self.pick_photos_dir).pack(side="left", padx=5)

        tk.Label(left_f, text="(Note: Name photos exactly as Enrollment No. e.g. ENR101.jpg)", font=("Segoe UI", 7, "italic"), fg=WARNING, bg=BG_CARD).pack(anchor="w", pady=(0, 15))

        right_f = self.create_card_frame(studio); right_f.pack(side="right", fill="both", expand=True, padx=(10,0)); right_f.config(padx=20, pady=20)
        tk.Label(right_f, text="Generation Scope", font=("Segoe UI", 11, "bold"), fg=SUCCESS, bg=BG_CARD).pack(anchor="w", pady=(0, 15))

        self.scope_var = tk.StringVar(value="all")
        tk.Radiobutton(right_f, text="All Registered Students", variable=self.scope_var, value="all", font=("Segoe UI", 9), fg=FG_PRIMARY, bg=BG_CARD, selectcolor=BG_MAIN, activebackground=BG_CARD, activeforeground=FG_PRIMARY).pack(anchor="w", pady=5)
        tk.Radiobutton(right_f, text="Filter by Class:", variable=self.scope_var, value="class", font=("Segoe UI", 9), fg=FG_PRIMARY, bg=BG_CARD, selectcolor=BG_MAIN, activebackground=BG_CARD, activeforeground=FG_PRIMARY).pack(anchor="w", pady=5)
        
        av_cls = sorted(list(set([s[6] for s in self.db.get_all_students()])))
        self.cb_id_class = ttk.Combobox(right_f, values=av_cls if av_cls else ["Play"], state="readonly", width=20)
        self.cb_id_class.set(av_cls[0] if av_cls else "Play"); self.cb_id_class.pack(anchor="w", padx=25, pady=(0, 15))

        tk.Radiobutton(right_f, text="Single Enrollment Number:", variable=self.scope_var, value="single", font=("Segoe UI", 9), fg=FG_PRIMARY, bg=BG_CARD, selectcolor=BG_MAIN, activebackground=BG_CARD, activeforeground=FG_PRIMARY).pack(anchor="w", pady=5)
        self.ent_id_single = tk.Entry(right_f, width=22, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid")
        self.ent_id_single.pack(anchor="w", padx=25, pady=(0, 25))

        self.create_btn(right_f, "Generate Exact Replica PDF", SUCCESS, self.process_id_cards, "🖨️").pack(anchor="w")

    def pick_logo(self):
        f = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")])
        if f: self.id_logo_path = f; self.lbl_logo.config(text=os.path.basename(f))

    def pick_sign(self):
        f = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")])
        if f: self.id_sign_path = f; self.lbl_sign.config(text=os.path.basename(f))
    
    def pick_photos_dir(self):
        d = filedialog.askdirectory()
        if d: self.id_photos_dir = d; self.lbl_photos.config(text=os.path.basename(d))

    def process_id_cards(self):
        scope = self.scope_var.get()
        students_to_print = []

        if scope == "all":
            for s in self.db.get_all_students():
                full_st = self.db.get_student(s[0])
                if full_st: students_to_print.append(full_st)
        elif scope == "class":
            target = self.cb_id_class.get()
            for s in self.db.get_all_students():
                full_st = self.db.get_student(s[0])
                if full_st and full_st[6] == target: students_to_print.append(full_st)
        elif scope == "single":
            roll = self.ent_id_single.get().strip()
            if not roll: return messagebox.showwarning("Input Needed", "Enter an Enrollment Number.")
            full_st = self.db.get_student(roll)
            if full_st: students_to_print.append(full_st)

        if not students_to_print: return messagebox.showinfo("No Records", "No matching students found.")

        try:
            pdf_path = generate_advanced_id_cards(students_to_print, logo_path=self.id_logo_path, photos_dir=self.id_photos_dir, sign_path=self.id_sign_path)
            messagebox.showinfo("Success", f"Generated {len(students_to_print)} exact-match cards.\nSaved to: {pdf_path}")
            if hasattr(os, 'startfile'): os.startfile(pdf_path)
            else: webbrowser.open(pdf_path)
        except Exception as e:
            messagebox.showerror("PDF Error", f"Failed to build ID cards: {e}")

    # --- VIEW: REGISTER (ADMISSION DESK FORM) ---
    def show_register(self):
        self.set_active_nav("register"); self.clear_content()
        tk.Label(self.content_frame, text="New Student Admission Desk", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")
        
        canvas_container = tk.Canvas(self.content_frame, bg=BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.content_frame, orient="vertical", command=canvas_container.yview)
        form_body = tk.Frame(canvas_container, bg=BG_CARD, padx=30, pady=25)

        canvas_window = canvas_container.create_window((0, 0), window=form_body, anchor="nw")

        def configure_form_body(event):
            canvas_container.configure(scrollregion=canvas_container.bbox("all"))
        def configure_canvas(event):
            canvas_container.itemconfig(canvas_window, width=event.width)

        form_body.bind("<Configure>", configure_form_body)
        canvas_container.bind("<Configure>", configure_canvas)
        canvas_container.configure(yscrollcommand=scrollbar.set)

        canvas_container.pack(side="left", fill="both", expand=True, pady=15)
        scrollbar.pack(side="right", fill="y", pady=15)

        tk.Label(form_body, text="STUDENT ADMISSION PROFILE", font=("Segoe UI", 12, "bold"), fg=ACCENT_PRI, bg=BG_CARD).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 15))

        labels = [
            ("Enrollment No (Unique) *:", 1, 0), ("Class Roll No *:", 1, 2),
            ("Form No.:", 2, 0), ("Admission No.:", 2, 2),
            ("Academic Year:", 3, 0), ("Student Name *:", 3, 2),
            ("Class / Grade *:", 4, 0), ("D.O.B (DD-MM-YYYY) *:", 4, 2),
            ("Aadhaar No.:", 5, 0), ("Place of Birth:", 5, 2),
            ("State:", 6, 0), ("Nationality:", 6, 2),
            ("Religion:", 7, 0), ("Gender:", 7, 2),
            ("Caste (General/SC/ST/OBC):", 8, 0), ("Father's Name:", 8, 2),
            ("Father's Occupation:", 9, 0), ("Mother's Name:", 9, 2),
            ("Mother's Occupation:", 10, 0), ("Residential Address:", 10, 2),
            ("Pin Code:", 11, 0), ("Mobile (Father's) *:", 11, 2),
            ("Mobile (Mother's):", 12, 0), ("E-mail ID:", 12, 2),
            ("Mother Tongue:", 13, 0), ("Blood Group:", 13, 2),
            ("Identification Marks:", 14, 0), ("Name of Previous School:", 14, 2),
            ("Name of Board:", 15, 0), ("% of Marks:", 15, 2),
            ("Reference / Centre:", 16, 0), ("Behaviour:", 16, 2),
            ("Transport Opted (Yes/No):", 17, 0), ("Ledger Folio No.:", 17, 2),
            ("Base Tuition Fee (₹):", 18, 0), ("Base Transport Fee (₹):", 18, 2)
        ]

        self.reg_entries = {}
        for text, r, c in labels:
            tk.Label(form_body, text=text, font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=r, column=c, sticky="w", padx=10, pady=6)
            if "Transport Opted" in text: 
                cb = ttk.Combobox(form_body, values=["No", "Yes"], state="readonly", width=22); cb.set("No"); cb.grid(row=r, column=c+1, padx=10, pady=6); self.reg_entries[text] = cb
            else: 
                e = tk.Entry(form_body, width=24, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid")
                e.grid(row=r, column=c+1, padx=10, pady=6)
                if "Academic Year" in text: e.insert(0, "2026-27")
                elif "Nationality" in text: e.insert(0, "Indian")
                elif "Mother Tongue" in text: e.insert(0, "Hindi")
                elif "Behaviour" in text: e.insert(0, "Normal")
                self.reg_entries[text] = e

        self.create_btn(form_body, "Save Admission & Print Form", SUCCESS, self.handle_admission_save, "🖨️").grid(row=19, column=0, columnspan=4, pady=(20, 10))

    def handle_admission_save(self):
        enr = self.reg_entries["Enrollment No (Unique) *:"].get().strip()
        roll = self.reg_entries["Class Roll No *:"].get().strip()
        name = self.reg_entries["Student Name *:"].get().strip()
        cls = self.reg_entries["Class / Grade *:"].get().strip()
        dob = self.reg_entries["D.O.B (DD-MM-YYYY) *:"].get().strip()
        contact = self.reg_entries["Mobile (Father's) *:"].get().strip()

        if not (enr and roll and name and cls and contact and dob):
            return messagebox.showerror("Error", "Enrollment No, Class Roll, Name, Class, D.O.B, and Contact are mandatory.")

        try:
            tuition = float(self.reg_entries["Base Tuition Fee (₹):"].get().strip() or 0)
            tr_fee = float(self.reg_entries["Base Transport Fee (₹):"].get().strip() or 0)
        except ValueError:
            return messagebox.showerror("Error", "Fees must be numeric.")

        data = (
            enr,
            self.reg_entries["Form No.:"].get().strip(),
            self.reg_entries["Admission No.:"].get().strip(),
            self.reg_entries["Academic Year:"].get().strip(),
            roll, name, cls, dob,
            self.reg_entries["Aadhaar No.:"].get().strip(),
            self.reg_entries["Place of Birth:"].get().strip(),
            self.reg_entries["State:"].get().strip(),
            self.reg_entries["Nationality:"].get().strip(),
            self.reg_entries["Religion:"].get().strip(),
            self.reg_entries["Gender:"].get().strip(),
            self.reg_entries["Caste (General/SC/ST/OBC):"].get().strip(),
            self.reg_entries["Father's Name:"].get().strip(),
            self.reg_entries["Father's Occupation:"].get().strip(),
            self.reg_entries["Mother's Name:"].get().strip(),
            self.reg_entries["Mother's Occupation:"].get().strip(),
            self.reg_entries["Residential Address:"].get().strip(),
            self.reg_entries["Pin Code:"].get().strip(),
            contact,
            self.reg_entries["Mobile (Mother's):"].get().strip(),
            self.reg_entries["E-mail ID:"].get().strip(),
            self.reg_entries["Mother Tongue:"].get().strip(),
            self.reg_entries["Blood Group:"].get().strip(),
            self.reg_entries["Identification Marks:"].get().strip(),
            self.reg_entries["Name of Previous School:"].get().strip(),
            self.reg_entries["Name of Board:"].get().strip(),
            self.reg_entries["% of Marks:"].get().strip(),
            self.reg_entries["Reference / Centre:"].get().strip(),
            self.reg_entries["Behaviour:"].get().strip(),
            self.reg_entries["Transport Opted (Yes/No):"].get().strip(),
            self.reg_entries["Ledger Folio No.:"].get().strip(),
            tuition, tr_fee
        )

        try:
            self.db.add_student(data)
            st_record = self.db.get_student(enr)
            pdf_path = generate_admission_form_pdf(st_record)
            messagebox.showinfo("Success", f"'{name}' successfully admitted!\n\nAdmission form generated at:\n{pdf_path}")
            if hasattr(os, 'startfile'): os.startfile(pdf_path)
            else: webbrowser.open(pdf_path)
            self.show_students()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", f"Enrollment No '{enr}' already exists in database.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to complete admission: {e}")

    # --- VIEW: BILLS & PAYMENTS ---
    def show_bills(self):
        self.set_active_nav("bills"); self.clear_content()
        tk.Label(self.content_frame, text="Bills & Payments Hub", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")

        bar = self.create_card_frame(self.content_frame); bar.pack(fill="x", pady=15); bar.config(padx=15, pady=12)
        tk.Label(bar, text="Filter:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).pack(side="left", padx=5)
        self.cb_status_filter = ttk.Combobox(bar, values=["All Status", "Unpaid", "Partially Paid", "Paid", "Cancelled"], state="readonly", width=14); self.cb_status_filter.set("All Status"); self.cb_status_filter.pack(side="left", padx=5); self.cb_status_filter.bind("<<ComboboxSelected>>", lambda e: self.populate_bills_tree())
        self.ent_bill_search = tk.Entry(bar, width=15, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); self.ent_bill_search.pack(side="left", padx=10); self.ent_bill_search.bind("<KeyRelease>", lambda e: self.populate_bills_tree())

        self.create_btn(bar, "WhatsApp", WA_GREEN, self.share_bill_whatsapp, "💬").pack(side="right", padx=3)
        self.create_btn(bar, "Receipts", PURPLE_ACC, self.open_receipts_history_dialog, "🧾").pack(side="right", padx=3)
        self.create_btn(bar, "View Bill", ACCENT_SEC, self.view_selected_bill_pdf, "📄").pack(side="right", padx=3)
        self.create_btn(bar, "Payment", SUCCESS, self.open_record_payment_dialog, "💵").pack(side="right", padx=3)
        self.create_btn(bar, "New Bill", ACCENT_PRI, self.open_generate_bill_dialog, "+").pack(side="right", padx=3)

        tbl_frame = tk.Frame(self.content_frame, bg=BG_MAIN); tbl_frame.pack(fill="both", expand=True)
        cols = ("bill_no", "enr_no", "name", "month", "total", "paid", "dues", "status", "due_date")
        self.tree_bills = ttk.Treeview(tbl_frame, columns=cols, show="headings", selectmode="browse")
        w_dict = {"bill_no":150, "enr_no":90, "name":180, "month":100, "total":100, "paid":100, "dues":100, "status":110, "due_date":100}
        for c, h in zip(cols, ["Bill No.", "Enrollment No", "Student Name", "Month", "Amount (₹)", "Paid (₹)", "Dues (₹)", "Status", "Due Date"]): self.tree_bills.heading(c, text=h); self.tree_bills.column(c, width=w_dict[c], minwidth=w_dict[c], anchor="center" if c not in ["name", "total", "paid", "dues"] else ("w" if c == "name" else "e"))
        scroll_y = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.tree_bills.yview); scroll_x = ttk.Scrollbar(tbl_frame, orient="horizontal", command=self.tree_bills.xview)
        self.tree_bills.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set); scroll_x.pack(side="bottom", fill="x"); self.tree_bills.pack(side="left", fill="both", expand=True); scroll_y.pack(side="right", fill="y")
        self.populate_bills_tree()

    def populate_bills_tree(self):
        for item in self.tree_bills.get_children(): self.tree_bills.delete(item)
        for r in self.db.get_all_bills(self.cb_status_filter.get(), self.ent_bill_search.get().strip()): self.tree_bills.insert("", "end", values=(r[0], r[1], r[2], r[3], f"{r[4]:.2f}", f"{r[5]:.2f}", f"{r[6]:.2f}", r[7], r[8]))

    def open_generate_bill_dialog(self):
        win = tk.Toplevel(self); win.title("Generate Demand Bill"); win.geometry("680x620"); win.configure(bg=BG_MAIN)
        form = self.create_card_frame(win); form.pack(fill="both", expand=True, padx=25, pady=25); form.config(padx=25, pady=20)
        tk.Label(form, text="Enrollment No *:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=1, column=0, sticky="w", pady=5)
        ent_roll = tk.Entry(form, width=20, bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); ent_roll.grid(row=1, column=1, pady=5)
        lbl_info = tk.Label(form, text="", font=("Segoe UI", 8, "italic"), fg=SUCCESS, bg=BG_CARD); lbl_info.grid(row=2, column=0, columnspan=3, sticky="w", pady=2)

        def fetch_fee():
            st = self.db.get_student(ent_roll.get().strip())
            if st:
                lbl_info.config(text=f"Selected: {st[5]} | Tuition: ₹{st[34]} | Transport: ₹{st[35]}")
                if "Monthly / Tuition Fee" in item_entries: item_entries["Monthly / Tuition Fee"].delete(0, tk.END); item_entries["Monthly / Tuition Fee"].insert(0, str(st[34]))
                if "Transport Charge" in item_entries and st[32] == "Yes": item_entries["Transport Charge"].delete(0, tk.END); item_entries["Transport Charge"].insert(0, str(st[35]))
            else: lbl_info.config(text="Student not found.", fg=DANGER)

        self.create_btn(form, "Fetch", ACCENT_SEC, fetch_fee).grid(row=1, column=2, padx=10)
        tk.Label(form, text="Billing Month *:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=3, column=0, sticky="w", pady=5)
        cb_month = ttk.Combobox(form, values=["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"], state="readonly", width=18); cb_month.set("April"); cb_month.grid(row=3, column=1, pady=5)
        tk.Label(form, text="Due Date:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=4, column=0, sticky="w", pady=5)
        ent_due = tk.Entry(form, width=20, bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); ent_due.insert(0, datetime.now().strftime("%Y-%m-10")); ent_due.grid(row=4, column=1, pady=5)

        fee_frame = tk.LabelFrame(form, text=" Fee Particulars Breakdown ", bg=BG_CARD, fg=FG_MUTED, padx=15, pady=15, bd=1); fee_frame.grid(row=5, column=0, columnspan=3, sticky="ew", pady=15)
        item_entries = {}
        for idx, head in enumerate(self.fee_heads):
            r, c = idx // 2, (idx % 2) * 2
            tk.Label(fee_frame, text=f"{head}:", font=("Segoe UI", 8), fg=FG_PRIMARY, bg=BG_CARD).grid(row=r, column=c, sticky="w", padx=5, pady=4)
            e = tk.Entry(fee_frame, width=12, bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); e.insert(0, "0"); e.grid(row=r, column=c+1, padx=10, pady=4); item_entries[head] = e

        def save_and_make_bill():
            enr = ent_roll.get().strip()
            st = self.db.get_student(enr)
            if not st: return messagebox.showerror("Error", "Valid Enrollment Number required.")
            month, due = cb_month.get(), ent_due.get().strip()
            bill_date, bill_no = datetime.now().strftime("%Y-%m-%d"), f"BILL-{datetime.now().strftime('%Y%m')}-{enr}-{datetime.now().strftime('%S')}"
            breakup_list, breakup_str_list, total_bill = [], [], 0.0
            for head, ent in item_entries.items():
                try:
                    val = float(ent.get().strip() or 0)
                    if val > 0: total_bill += val; breakup_list.append((head, val)); breakup_str_list.append(f"{head}: ₹{val:,.0f}")
                except ValueError: return messagebox.showerror("Error", f"Invalid amount for {head}")
            if total_bill <= 0: return messagebox.showerror("Error", "Bill amount must be > 0.")
            self.db.create_bill(bill_no, enr, month, bill_date, due, total_bill, ", ".join(breakup_str_list))
            try:
                pdf_path = generate_demand_bill_pdf((bill_no, enr, month, bill_date, due, total_bill, 0, 'Unpaid', ", ".join(breakup_str_list)), st, breakup_list)
                if hasattr(os, 'startfile'): os.startfile(pdf_path)
                else: webbrowser.open(pdf_path)
            except Exception as e: messagebox.showerror("File Error", f"Could not generate PDF: {e}")
            self.populate_bills_tree(); win.destroy()

        self.create_btn(form, "Generate & Print Bill", ACCENT_PRI, save_and_make_bill, "📄").grid(row=6, column=0, columnspan=3, pady=15)

    def open_record_payment_dialog(self):
        sel = self.tree_bills.selection()
        if not sel: return messagebox.showwarning("Select", "Select a bill to record payment against.")
        bill_no = self.tree_bills.item(sel[0])["values"][0]
        bill = self.db.get_bill_by_id(bill_no); st = self.db.get_student(bill[1])

        win = tk.Toplevel(self); win.title("Record Payment"); win.geometry("460x420"); win.configure(bg=BG_MAIN)
        f = self.create_card_frame(win); f.pack(fill="both", expand=True, padx=25, pady=25); f.config(padx=25, pady=20)
        dues = bill[5] - bill[6]
        tk.Label(f, text="PROCESS PAYMENT", font=("Segoe UI", 12, "bold"), fg=SUCCESS, bg=BG_CARD).grid(row=0, column=0, columnspan=2, pady=(0, 15))
        tk.Label(f, text=f"Student: {st[5]} (Bill: {bill_no})", font=("Segoe UI", 9), fg=FG_MUTED, bg=BG_CARD).grid(row=1, column=0, columnspan=2, sticky="w", pady=2)
        tk.Label(f, text=f"Total: ₹{bill[5]:,.2f}  |  Dues: ₹{dues:,.2f}", font=("Segoe UI", 10, "bold"), fg=WARNING, bg=BG_CARD).grid(row=2, column=0, columnspan=2, sticky="w", pady=(2,15))
        tk.Label(f, text="Amount Paying (₹) *:", font=("Segoe UI", 9, "bold"), fg=FG_PRIMARY, bg=BG_CARD).grid(row=3, column=0, sticky="w", pady=8)
        ent_amt = tk.Entry(f, width=20, bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); ent_amt.insert(0, str(dues)); ent_amt.grid(row=3, column=1, pady=8)
        tk.Label(f, text="Payment Mode:", font=("Segoe UI", 9, "bold"), fg=FG_PRIMARY, bg=BG_CARD).grid(row=4, column=0, sticky="w", pady=8)
        cb_mode = ttk.Combobox(f, values=["Cash", "UPI / QR", "Card", "Bank Transfer"], state="readonly", width=18); cb_mode.set("Cash"); cb_mode.grid(row=4, column=1, pady=8)
        tk.Label(f, text="Receipt No.:", font=("Segoe UI", 9, "bold"), fg=FG_PRIMARY, bg=BG_CARD).grid(row=5, column=0, sticky="w", pady=8)
        ent_rec = tk.Entry(f, width=20, bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); ent_rec.insert(0, f"REC-{datetime.now().strftime('%d%H%M')}"); ent_rec.grid(row=5, column=1, pady=8)

        def save_pay():
            try:
                amt = float(ent_amt.get().strip())
                if amt <= 0: raise ValueError
            except ValueError: return messagebox.showerror("Error", "Enter a valid positive number.")
            success, msg, payment_data = self.db.record_payment(bill_no, amt, cb_mode.get(), ent_rec.get().strip())
            if success:
                try:
                    pdf_path = generate_receipt_pdf(payment_data, bill, st)
                    messagebox.showinfo("Success", f"Payment recorded!\nReceipt saved to:\n{pdf_path}")
                    if hasattr(os, 'startfile'): os.startfile(pdf_path)
                except Exception as e: messagebox.showerror("PDF Error", f"Payment saved, but receipt failed:\n{e}")
                self.populate_bills_tree(); win.destroy()
            else: messagebox.showerror("Error", msg)

        self.create_btn(f, "Confirm & Generate Receipt", SUCCESS, save_pay, "✅").grid(row=6, column=0, columnspan=2, pady=25)

    def view_selected_bill_pdf(self):
        sel = self.tree_bills.selection()
        if not sel: return messagebox.showwarning("Select Bill", "Select a bill first.")
        bill_no = self.tree_bills.item(sel[0])["values"][0]
        bill = self.db.get_bill_by_id(bill_no)
        st = self.db.get_student(bill[1])
        items_list = []
        for part in bill[8].split(", "):
            if ":" in part: h, v = part.split(": ₹"); items_list.append((h, float(v.replace(",", ""))))
        try:
            pdf_path = generate_demand_bill_pdf(bill, st, items_list)
            if hasattr(os, 'startfile'): os.startfile(pdf_path)
            else: webbrowser.open(pdf_path)
        except PermissionError: messagebox.showerror("File Open", "Close the existing PDF before updating it.")

    def open_receipts_history_dialog(self):
        sel = self.tree_bills.selection()
        if not sel: return messagebox.showwarning("Select Bill", "Select a bill to view its receipts.")
        bill_no = self.tree_bills.item(sel[0])["values"][0]; payments = self.db.get_payments_for_bill(bill_no)
        if not payments: return messagebox.showinfo("No Receipts", "No payments have been recorded for this bill yet.")

        win = tk.Toplevel(self); win.title(f"Receipts - {bill_no}"); win.geometry("620x420"); win.configure(bg=BG_MAIN)
        f = self.create_card_frame(win); f.pack(fill="both", expand=True, padx=20, pady=20); f.config(padx=15, pady=15)
        tk.Label(f, text=f"Historical Receipts for {bill_no}", font=("Segoe UI", 12, "bold"), fg=PURPLE_ACC, bg=BG_CARD).pack(anchor="w", pady=(0, 10))
        cols = ("date", "mode", "amount", "receipt_no"); tree = ttk.Treeview(f, columns=cols, show="headings", height=8)
        for c, h in zip(cols, ["Date & Time", "Mode", "Amount Paid (₹)", "Receipt No."]): tree.heading(c, text=h); tree.column(c, anchor="center")
        for p in payments: tree.insert("", "end", values=(p[1], p[2], f"{p[3]:.2f}", p[4]))
        tree.pack(fill="both", expand=True, pady=10)

        def open_pdf():
            p_sel = tree.selection()
            if not p_sel: return messagebox.showwarning("Select Receipt", "Select a receipt row to open.")
            rec_no = tree.item(p_sel[0])["values"][3]; pdf_path = os.path.join(RECEIPTS_DIR, f"Receipt_{rec_no}.pdf")
            if os.path.exists(pdf_path): os.startfile(pdf_path) if hasattr(os, 'startfile') else webbrowser.open(pdf_path)
            else: messagebox.showerror("Missing File", "Receipt PDF file missing.")
        self.create_btn(f, "Re-Print Selected Receipt", SUCCESS, open_pdf, "📄").pack(pady=5)

    def share_bill_whatsapp(self):
        sel = self.tree_bills.selection()
        if not sel: return messagebox.showwarning("Select Bill", "Select a bill to share on WhatsApp.")
        bill_no = self.tree_bills.item(sel[0])["values"][0]
        bill, st = self.db.get_bill_by_id(bill_no), self.db.get_student(self.db.get_bill_by_id(bill_no)[1])
        phone_clean = re.sub(r'[^\d+]', '', st[21].strip()) 
        if len(phone_clean) == 10: phone_clean = "+91" + phone_clean
        elif len(phone_clean) > 10 and not phone_clean.startswith("+"): phone_clean = "+" + phone_clean
        msg = f"Dear Parent of {st[5]},\n\nReminder regarding fee bill for {bill[3]}.\n\n🔹 *Bill No:* {bill_no}\n🔹 *Pending Dues:* ₹{(bill[5]-bill[6]):,.2f}\n🔹 *Due Date:* {bill[4]}\n\nPlease find the PDF attached.\nRegards,\n*Intech Kids*"
        webbrowser.open(f"https://wa.me/{phone_clean.replace('+', '')}?text={urllib.parse.quote(msg)}")
        pdf_path = os.path.join(BILLS_DIR, f"DemandBill_{bill_no}.pdf")
        if not os.path.exists(pdf_path):
            items_list = []
            for part in bill[8].split(", "):
                if ":" in part: h, v = part.split(": ₹"); items_list.append((h, float(v.replace(",", ""))))
            pdf_path = generate_demand_bill_pdf(bill, st, items_list)
        try: subprocess.Popen(f'explorer /select,"{pdf_path}"')
        except: os.startfile(BILLS_DIR)

    # --- VIEW: FINANCE & EXPENSES ---
    def show_finance(self):
        self.set_active_nav("finance"); self.clear_content()
        tk.Label(self.content_frame, text="Finance & Expense Tracker", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")
        top_frame = tk.Frame(self.content_frame, bg=BG_MAIN); top_frame.pack(fill="x")
        
        form = self.create_card_frame(top_frame); form.pack(side="left", fill="both", expand=True, padx=(0, 10)); form.config(padx=20, pady=20)
        tk.Label(form, text="Record New Expense", font=("Segoe UI", 11, "bold"), fg=FINANCE_ACC, bg=BG_CARD).grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="w")
        tk.Label(form, text="Category:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=1, column=0, sticky="w", pady=5)
        cb_cat = ttk.Combobox(form, values=["Staff Salary", "Rent", "Electricity / Utilities", "Maintenance", "Events & Functions", "Miscellaneous"], state="readonly", width=22); cb_cat.set("Staff Salary"); cb_cat.grid(row=1, column=1, pady=5)
        tk.Label(form, text="Amount (₹):", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=2, column=0, sticky="w", pady=5)
        ent_amt = tk.Entry(form, width=24, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); ent_amt.grid(row=2, column=1, pady=5)
        tk.Label(form, text="Description:", font=("Segoe UI", 9, "bold"), fg=FG_MUTED, bg=BG_CARD).grid(row=3, column=0, sticky="w", pady=5)
        ent_desc = tk.Entry(form, width=24, font=("Segoe UI", 10), bg=BG_MAIN, fg=FG_PRIMARY, insertbackground=FG_PRIMARY, bd=1, relief="solid"); ent_desc.grid(row=3, column=1, pady=5)

        def save_exp():
            try:
                amt = float(ent_amt.get().strip())
                if amt <= 0: raise ValueError
            except ValueError: return messagebox.showerror("Error", "Enter a valid positive amount.")
            self.db.add_expense(cb_cat.get(), amt, ent_desc.get().strip(), self.current_user)
            messagebox.showinfo("Success", "Expense recorded."); self.show_finance()
        self.create_btn(form, "Save Expense", FINANCE_ACC, save_exp, "💾").grid(row=4, column=0, columnspan=2, pady=15)

        history_frame = self.create_card_frame(self.content_frame); history_frame.pack(fill="both", expand=True, pady=15)
        tk.Label(history_frame, text="Recent Expenses", font=("Segoe UI", 11, "bold"), fg=FG_PRIMARY, bg=BG_CARD).pack(anchor="w", padx=15, pady=10)
        cols = ("date", "cat", "amt", "desc", "user"); tree = ttk.Treeview(history_frame, columns=cols, show="headings", height=8)
        for c, h in zip(cols, ["Date", "Category", "Amount (₹)", "Description", "Recorded By"]): tree.heading(c, text=h); tree.column(c, anchor="center" if c != "desc" else "w")
        for r in self.db.get_all_expenses(): tree.insert("", "end", values=(r[1], r[2], f"{r[3]:.2f}", r[4], r[5]))
        tree.pack(fill="both", expand=True, padx=15, pady=(0,15))

    # --- VIEW: IMPORT EXCEL ---
    def show_import_excel(self):
        self.set_active_nav("import"); self.clear_content()
        tk.Label(self.content_frame, text="Bulk Import Students", font=("Segoe UI", 18, "bold"), fg=FG_PRIMARY, bg=BG_MAIN).pack(anchor="w")
        card = self.create_card_frame(self.content_frame); card.pack(fill="x", pady=20); card.config(padx=30, pady=30)
        tk.Label(card, text="Step 1: Download Format", font=("Segoe UI", 11, "bold"), fg=ACCENT_PRI, bg=BG_CARD).pack(anchor="w")
        self.create_btn(card, "Download Excel Template", ACCENT_SEC, self.download_excel_template, "📥").pack(anchor="w", pady=(10, 25))
        tk.Label(card, text="Step 2: Upload Data", font=("Segoe UI", 11, "bold"), fg=SUCCESS, bg=BG_CARD).pack(anchor="w")
        self.create_btn(card, "Select & Import Excel File", SUCCESS, self.execute_excel_import, "📂").pack(anchor="w", pady=10)

    def download_excel_template(self):
        dest = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Students_Import_Template.xlsx")
        if not dest: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Students"
        ws.append(["Enrollment No", "Form No", "Admission No", "Academic Year", "Class Roll", "Student Name", "Class", "D.O.B", "Aadhaar", "Place of Birth", "State", "Nationality", "Religion", "Gender", "Caste", "Father Name", "Father Occ", "Mother Name", "Mother Occ", "Address", "Pin", "Contact", "Mother Mob", "Email", "Mother Tongue", "Blood", "Marks", "Prev School", "Board", "%", "Ref", "Behaviour", "Transport", "Folio", "Tuition", "Transport Fee"])
        wb.save(dest); messagebox.showinfo("Saved", f"Template saved to:\n{dest}")

    def execute_excel_import(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx;*.xls")])
        if not file_path: return
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            rows = list(wb.active.iter_rows(values_only=True)); student_data = []
            for row in rows[1:]:
                if not row[0] or not row[5]: continue 
                enr = str(row[0]).strip()
                form_no = str(row[1]).strip() if len(row)>1 and row[1] else "N/A"
                adm_no = str(row[2]).strip() if len(row)>2 and row[2] else "N/A"
                acy = str(row[3]).strip() if len(row)>3 and row[3] else "2026-27"
                c_roll = str(row[4]).strip() if len(row)>4 and row[4] else "N/A"
                name = str(row[5]).strip()
                cls = str(row[6]).strip() if len(row)>6 and row[6] else "Play"
                dob = str(row[7]).strip() if len(row)>7 and row[7] else "N/A"
                aadhaar = str(row[8]).strip() if len(row)>8 and row[8] else "N/A"
                pob = str(row[9]).strip() if len(row)>9 and row[9] else "N/A"
                state = str(row[10]).strip() if len(row)>10 and row[10] else "N/A"
                nat = str(row[11]).strip() if len(row)>11 and row[11] else "Indian"
                rel = str(row[12]).strip() if len(row)>12 and row[12] else "N/A"
                gen = str(row[13]).strip() if len(row)>13 and row[13] else "N/A"
                caste = str(row[14]).strip() if len(row)>14 and row[14] else "General"
                f_name = str(row[15]).strip() if len(row)>15 and row[15] else "N/A"
                f_occ = str(row[16]).strip() if len(row)>16 and row[16] else "N/A"
                m_name = str(row[17]).strip() if len(row)>17 and row[17] else "N/A"
                m_occ = str(row[18]).strip() if len(row)>18 and row[18] else "N/A"
                addr = str(row[19]).strip() if len(row)>19 and row[19] else "N/A"
                pin = str(row[20]).strip() if len(row)>20 and row[20] else "N/A"
                cont = str(row[21]).strip() if len(row)>21 and row[21] else "N/A"
                m_mob = str(row[22]).strip() if len(row)>22 and row[22] else "N/A"
                email = str(row[23]).strip() if len(row)>23 and row[23] else "N/A"
                m_tongue = str(row[24]).strip() if len(row)>24 and row[24] else "Hindi"
                blood = str(row[25]).strip() if len(row)>25 and row[25] else "N/A"
                marks = str(row[26]).strip() if len(row)>26 and row[26] else "N/A"
                sch = str(row[27]).strip() if len(row)>27 and row[27] else "N/A"
                board = str(row[28]).strip() if len(row)>28 and row[28] else "N/A"
                perc = str(row[29]).strip() if len(row)>29 and row[29] else "N/A"
                ref = str(row[30]).strip() if len(row)>30 and row[30] else "N/A"
                behav = str(row[31]).strip() if len(row)>31 and row[31] else "Normal"
                trans = str(row[32]).strip() if len(row)>32 and row[32] else "No"
                folio = str(row[33]).strip() if len(row)>33 and row[33] else "N/A"
                tuit = float(row[34]) if len(row)>34 and row[34] else 0.0
                t_fee = float(row[35]) if len(row)>35 and row[35] else 0.0

                student_data.append((enr, form_no, adm_no, acy, c_roll, name, cls, dob, aadhaar, pob, state, nat, rel, gen, caste, f_name, f_occ, m_name, m_occ, addr, pin, cont, m_mob, email, m_tongue, blood, marks, sch, board, perc, ref, behav, trans, folio, tuit, t_fee))

            if not student_data: return
            ins, upd = self.db.bulk_import_students(student_data)
            messagebox.showinfo("Complete", f"Processed {len(student_data)} admission records.\nAdded: {ins}\nUpdated: {upd}")
        except Exception as e: messagebox.showerror("Error", str(e))

    # --- ADMIN UTILS ---
    def run_promotion(self):
        if messagebox.askyesno("Confirm Promotion", "This will shift all students to their next grade. Proceed?"):
            promoted = self.db.promote_all_students(self.current_user)
            messagebox.showinfo("Success", f"{promoted} students promoted to the next grade."); self.show_dashboard()

    def restore_backup_dialog(self):
        backup = filedialog.askopenfilename(initialdir=BACKUP_DIR, filetypes=[("DB", "*.db")])
        if backup and messagebox.askyesno("Confirm", "Replace current database?"):
            self.db.log_audit(self.current_user, "RESTORE", f"Restored DB from {os.path.basename(backup)}")
            self.db.close(); shutil.copy2(backup, self.db.db_path); self.db.reconnect()
            self.show_dashboard(); messagebox.showinfo("Restored", "Database restored.")


if __name__ == "__main__":
    app = LoginApp()
    app.mainloop()